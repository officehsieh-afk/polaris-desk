#!/usr/bin/env python3
"""從 R6 `Ontology_V1.xlsx` 重生 canonical seed CSV（單一事實來源）。

`migrations/2026-06-18_create_r6_ontology.sql` 以 `docs/r6/ontology/seeds/*.csv`
為唯一來源把 ontology 落地進 `polaris_core`。ontology 分頁異動時（例如 PR #104
改了 04_disclosure_events / 10_news_source_whitelist），請用本腳本重生對應 seed，
**勿手改單列**，以免 CSV 與 xlsx 不一致。

用法（從 repo 根目錄）：
    # 重生全部已對照的分頁
    python scripts/gen_ontology_seeds.py
    # 只重生指定 seed（#104 受影響的兩張）
    python scripts/gen_ontology_seeds.py --only disclosure_event news_source_whitelist
    # 只印不寫（檢視 diff 用）
    python scripts/gen_ontology_seeds.py --only news_source_whitelist --dry-run

只讀 xlsx、只寫 repo 內 seeds/，不碰 BigQuery。
"""
from __future__ import annotations

import argparse
import csv
import io
import sys
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
XLSX = REPO / "docs/r6/ontology/Ontology_V1.xlsx"
SEEDS = REPO / "docs/r6/ontology/seeds"

# seed 名 → xlsx 分頁名（分頁名含中文後綴；對照 migration 註解的 sheet→seed→table）
SHEET_BY_SEED = {
    "industry": "01_industry產業",
    "financial_metric": "03_financial_metric財務指標",
    "disclosure_event": "04_disclosure_events事件",
    "compliance_term": "05_compliance_terms法遵名詞",
    "theme": "06_theme",
    "risk_signal": "07_risk_signal",
    "company_theme_map": "08_company_theme_mapping",
    "company_industry_map": "09_company_industry_mapping",
    "news_source_whitelist": "10_news_source_whitelist",
    "source_taxonomy_map": "12_source_taxonomy_mapping",
    "revenue_field": "14_revenue_metrics_extension",
    "quarter": "16_quarter_table",
}


def _rows(ws):
    """回傳 [header, *data]；於第一個全空列截斷（去掉尾端空列）。"""
    out: list[list[str]] = []
    for r in ws.iter_rows(values_only=True):
        if all(c is None or str(c).strip() == "" for c in r):
            break
        out.append(["" if c is None else str(c).strip() for c in r])
    return out


def gen_one(seed: str, sheet: str, *, dry_run: bool) -> int:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    if sheet not in wb.sheetnames:
        sys.exit(f"分頁不存在：{sheet}（有：{wb.sheetnames}）")
    rows = _rows(wb[sheet])
    if not rows:
        sys.exit(f"分頁 {sheet} 沒有資料列")

    buf = io.StringIO()
    w = csv.writer(buf, lineterminator="\n")
    w.writerows(rows)
    text = buf.getvalue()

    n_data = len(rows) - 1
    if dry_run:
        print(f"# {seed}.csv ← {sheet}  ({n_data} 資料列, {len(rows[0])} 欄)")
        print(text)
    else:
        (SEEDS / f"{seed}.csv").write_text(text, encoding="utf-8")
        print(f"✓ {seed}.csv  ({n_data} 資料列, {len(rows[0])} 欄)  ← {sheet}")
    return n_data


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--only", nargs="*", help="只處理這些 seed（預設全部）")
    ap.add_argument("--dry-run", action="store_true", help="只印不寫")
    args = ap.parse_args()

    if sys.stdout.encoding != "utf-8":
        sys.stdout.reconfigure(encoding="utf-8")

    targets = args.only or list(SHEET_BY_SEED)
    unknown = [t for t in targets if t not in SHEET_BY_SEED]
    if unknown:
        sys.exit(f"未知 seed：{unknown}（可選：{list(SHEET_BY_SEED)}）")

    for seed in targets:
        gen_one(seed, SHEET_BY_SEED[seed], dry_run=args.dry_run)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
